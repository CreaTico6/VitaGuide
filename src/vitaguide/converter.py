from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


def _slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch)).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "item"


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_for_match(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch)).lower()
    return re.sub(r"\s+", " ", text).strip()


def _split_entities(raw: str) -> List[str]:
    if not raw:
        return []
    text = re.sub(r"\b\d+(?:-\d+)?\b\.?", "", raw)
    text = text.replace(";", ",")
    parts = [p.strip(" .") for p in text.split(",")]
    clean = []
    seen = set()
    for part in parts:
        if not part:
            continue
        if len(part) < 3:
            continue
        key = unicodedata.normalize("NFKD", part).encode("ascii", "ignore").decode("ascii").lower()
        if key in seen:
            continue
        seen.add(key)
        clean.append(part)
    return clean


def _find_column(headers: Dict[str, int], patterns: List[str]) -> int:
    normalized_headers = { _normalize_for_match(header): idx for header, idx in headers.items() }
    for pattern in patterns:
        normalized_pattern = _normalize_for_match(pattern)
        for header, idx in normalized_headers.items():
            if normalized_pattern in header:
                return idx
    return -1


def _sheet_records(ws, item_type: str) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = [str(c or "").strip().lower() for c in rows[0]]
    headers = {h: i for i, h in enumerate(header_row) if h}

    name_idx = _find_column(headers, ["suplemento", "fitoterap"])
    interaction_idx = _find_column(headers, ["intera", "farmacos", "farmacos intervenientes"])
    contraindication_idx = _find_column(headers, ["contraindica", "precau"])
    adverse_idx = _find_column(headers, ["efeitos", "limites", "riscos"])

    records: List[Dict[str, Any]] = []
    for line_no, row in enumerate(rows[1:], start=2):
        if name_idx < 0 or name_idx >= len(row):
            continue
        name = _clean_text(row[name_idx])
        if not name:
            continue

        interactions = _clean_text(row[interaction_idx]) if interaction_idx >= 0 and interaction_idx < len(row) else ""
        contraindications = _clean_text(row[contraindication_idx]) if contraindication_idx >= 0 and contraindication_idx < len(row) else ""
        adverse = _clean_text(row[adverse_idx]) if adverse_idx >= 0 and adverse_idx < len(row) else ""

        records.append(
            {
                "id": _slugify(name),
                "name": name,
                "aliases": [],
                "type": item_type,
                "interactions_text": interactions,
                "contraindications_text": contraindications,
                "adverse_text": adverse,
                "external_entities": _split_entities(interactions),
                "source": {"sheet": ws.title, "line": line_no},
            }
        )
    return records


def build_dataset(xlsx_path: Path) -> Dict[str, Any]:
    wb = load_workbook(str(xlsx_path), data_only=True)
    sheets = wb.worksheets
    if not sheets:
        raise ValueError("Workbook without sheets")

    items: List[Dict[str, Any]] = []
    if len(sheets) >= 1:
        items.extend(_sheet_records(sheets[0], "suplemento"))
    if len(sheets) >= 2:
        items.extend(_sheet_records(sheets[1], "fitoterapico"))

    external_catalog = sorted(
        {entity for item in items for entity in item.get("external_entities", [])},
        key=lambda s: s.lower(),
    )

    return {
        "version": 1,
        "source_file": str(xlsx_path.name),
        "items": items,
        "external_catalog": external_catalog,
    }


def write_dataset(xlsx_path: Path, output_path: Path) -> None:
    data = build_dataset(xlsx_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
